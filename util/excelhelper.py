import json
import logging
import os

from openpyxl import Workbook

from util import toolkit

logging.basicConfig(filename='info.log', level=logging.DEBUG)


def json_to_list(job_type_json_dir):
    job_type_lists = []  # 每一个特定职位的list
    for each_json in os.listdir(job_type_json_dir):
        with open(job_type_json_dir + '/' + each_json, 'r', encoding='utf-8') as f:
            for each_line in f.readlines():
                json_content = "{'joblist':" + each_line + "}"
                st_json_str = json_content.replace("\'", '\"').replace('None', 'null').replace('False', 'false')
                # st_json_str = ast.literal_eval(json_content)
                print(st_json_str)
                try:
                    json_obj = json.loads(st_json_str)
                except:
                    pass
                job_type_lists.append(json_obj)

    return job_type_lists


def write_excel(output_path, lists, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "职位信息"
    ws.cell(row=1, column=1).value = '发布时间'
    ws.cell(row=1, column=2).value = '工作时间'
    ws.cell(row=1, column=3).value = '职位名称'
    ws.cell(row=1, column=4).value = '公司名称'
    ws.cell(row=1, column=5).value = '公司规模'
    ws.cell(row=1, column=6).value = '所在城市'
    ws.cell(row=1, column=7).value = '学历要求'
    ws.cell(row=1, column=8).value = '工作经验'
    ws.cell(row=1, column=9).value = '职位薪酬'
    ws.cell(row=1, column=10).value = '平均薪资'

    rownum = 2

    for each_item in lists:
        info_list = each_item.get('joblist')
        for each_job_info_obj in info_list:
            ws.cell(row=rownum, column=1).value = each_job_info_obj['formatCreateTime']
            ws.cell(row=rownum, column=2).value = each_job_info_obj['workYear']
            ws.cell(row=rownum, column=3).value = each_job_info_obj['positionName']
            ws.cell(row=rownum, column=4).value = each_job_info_obj['companyFullName']
            ws.cell(row=rownum, column=5).value = each_job_info_obj['companySize']
            ws.cell(row=rownum, column=6).value = each_job_info_obj['city']
            ws.cell(row=rownum, column=7).value = each_job_info_obj['education']
            ws.cell(row=rownum, column=8).value = each_job_info_obj['workYear']
            ws.cell(row=rownum, column=9).value = each_job_info_obj['salary']
            ws.cell(row=rownum, column=10).value = toolkit.normalize(each_job_info_obj['salary'])
            rownum += 1
    wb.save(output_path + filename + '.xlsx')
    print('Excel生成成功!')


def write_csv(output_path, lists, filename):
    f = open(output_path  + filename + '.csv', 'wt',
             encoding='utf-8')
    f.write('发布时间,工作时间,职位名称,公司名称,公司规模,所在城市,学历要求,工作经验,单位薪酬,平均薪资')
    f.write('\n')
    for each_item in lists:
        info_list = each_item.get('joblist')
        for each_job_info_obj in info_list:
            f.write(each_job_info_obj['formatCreateTime'] + ',' + each_job_info_obj['workYear'] + ','
                    + each_job_info_obj['positionName'] + ',' + str(each_job_info_obj['positionId']) + ','
                    + str(each_job_info_obj['companyId']) + ',' + each_job_info_obj['positionAdvantage'] + ','
                    + each_job_info_obj['companyFullName'] + ',' + each_job_info_obj['city'] + ','
                    + each_job_info_obj['education'] + ',' + each_job_info_obj['industryField'] + ','
                    + each_job_info_obj['financeStage'] + ',' + each_job_info_obj['salary'] + ','
                    + str(each_job_info_obj['companySize']) + ',' + str(toolkit.normalize(each_job_info_obj['salary'])))
            f.write('\n')
        f.flush()
    f.close()


def process_json(json_file_path, output_path):
    if os.path.exists(json_file_path):
        dir_list = os.listdir(json_file_path)
        for each_dir in dir_list:
            print(json_file_path + os.path.sep + each_dir)
            lists = json_to_list(json_file_path + os.path.sep + each_dir)
            write_excel(output_path, lists, each_dir)


if __name__ == '__main__':
    logging.info('start generating Excel file...')
    # process('/Users/iceke/PycharmProjects/LagouJob/data/')
    logging.info('Done! Please check your result...')
